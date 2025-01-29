import pytest
import openpyxl
class home1:
    test_homepage_data = [{"first_name":"Nesar", "mail":"nesar@gmail.com", "password":"123456", "gender":"Male"},
                          {"first_name":"XYZ", "mail":"xyz@gmail.com", "password":"123456", "gender":"Female"}]

    @staticmethod
    def get_data():
        wb = openpyxl.load_workbook("A:\\Python\\Pycharm\\Python Selenium\\PythonSeleniumProject1\\Udemy_Framework\\Testdata\\Book.xlsx")
        sheet = wb.active
        Dict = {}
        for i in range(2, sheet.max_row + 1):
            for j in range(1, sheet.max_column + 1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            return [Dict]