import openpyxl
wb = openpyxl.load_workbook("A:\\Python\\Pycharm\\Python Selenium\\PythonSeleniumProject1\\Udemy_Framework\\Testdata\\Book.xlsx")
sheet = wb.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)

print(sheet.max_row)
print(sheet.max_column)
# print(sheet["A2"].value)

for i in range(2, sheet.max_row + 1):
    # if sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(1, sheet.max_column + 1):
            Dict[sheet.cell(row= 1, column= j).value] = sheet.cell(row= i, column= j).value
        print(Dict)
